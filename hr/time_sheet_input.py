
import streamlit as st
from openpyxl import load_workbook
from datetime import datetime
import os

# مسار الملف
file_path = r"C:\Users\Eng.Madyan\OneDrive\Desktop\قبل العيد\version-2\D_To_D sheet-DESKTOP-8NH5M00.xlsx"
sheet_name = "التشغيل"

# واجهة الإدخال
st.title("📝 تسجيل الحضور والانصراف")
name = st.text_input("👤 الاسم")
date_input = st.date_input("📅 التاريخ", datetime.today())

if st.button("➕ إضافة صف جديد"):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # نحدد أول صف فاضي
    row = 159
    while ws[f"R{row}"].value:
        row += 1

    # نكتب البيانات
    now = datetime.now()
    ws[f"R{row}"] = date_input
    ws[f"S{row}"] = ""
    ws[f"T{row}"] = now.strftime("%Y-%m-%d %H:%M:%S")  # وقت الدخول
    ws[f"U{row}"] = ""
    ws[f"V{row}"] = name

    wb.save(file_path)
    st.success(f"✅ تم إضافة الصف رقم {row}")

if st.button("📤 تسجيل وقت الخروج"):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    found = False
    today_str = date_input.strftime("%Y-%m-%d")

    for row in range(4, ws.max_row + 1):
        row_date = ws[f"R{row}"].value
        row_name = ws[f"V{row}"].value

        # التأكد من تطابق الاسم والتاريخ (باليوم فقط)
        if row_name == name and row_date and row_date.strftime("%Y-%m-%d") == today_str:
            ws[f"U{row}"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # وقت الخروج
            found = True
            break

    if found:
        wb.save(file_path)
        st.success("✅ تم تسجيل وقت الخروج بنجاح")
    else:
        st.warning("⚠️ لم يتم العثور على صف مطابق للاسم والتاريخ")


